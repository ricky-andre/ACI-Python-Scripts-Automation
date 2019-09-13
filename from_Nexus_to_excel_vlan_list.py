import openpyxl
import re
import os.path

# directory where the excel file will be created
excel_dir = "C:/Users/my_output_directory/"
# excel filename, will be created in the above directory
file_name = "ACI_vlan_list.xlsx"

# directory where Nexus configurations will be read
config_dir = "C:/Users/customer_configurations_path/"
# Nexus configuration filenames
config_files = ["Nexus_config_1.txt", "Nexus_config_2.txt", "Nexus_config_3.txt"]

if not os.path.isfile(excel_dir + file_name):
    targ = openpyxl.Workbook()
else:
    targ = openpyxl.load_workbook(excel_dir + file_name)

if 'Network' in targ.sheetnames:
    sheet = targ.get_sheet_by_name("Network")
else:
    sheet = targ.create_sheet("Network")

VLAN = 0
VRF = ""
IP_ADDRESS = ""
description = ""
HSRP_GROUP = ""
HSRP = ""
if_list = {}

vrf = ""
counter = 1
sheet['A'+str(counter)].value = "Apparato"
sheet['B'+str(counter)].value = "vrf"
sheet['C'+str(counter)].value = "vlan #"
sheet['D'+str(counter)].value = "l2 vlan name"
sheet['E'+str(counter)].value = "ip address"
sheet['F'+str(counter)].value = "hsrp grp"
sheet['G'+str(counter)].value = "hsrp vip"
sheet['H'+str(counter)].value = "description"

# L2_ONLY vlans are NOT printed in this
for read_file in config_files:
    if_list[read_file] = {}
    l2_vlan_names = {}
    l2_only_vlans = {}
    # list of temporary variables to save all data
    VLAN = 0
    VRF = ''
    PORT_MODE = ''
    TRUNK = ''
    IP_ADDRESS = ''
    description = ''
    L2_VLAN = 0
    HSRP_GROUP = ''
    HSRP = ''
    config = open(config_dir + read_file,"r")
    for line in config:
        if (re.search("^vlan (\d+)", line)):
            L2_VLAN = re.search("^vlan (\d+)", line).group(1)
            # print("inside vlan "+L2_VLAN)
        elif (re.search("^\s+name (.*)\s", line) and L2_VLAN):
            l2_vlan_names[L2_VLAN] = re.search("^\s+name (.*)\s", line).group(1)
            l2_only_vlans[L2_VLAN] = re.search("^\s+name (.*)\s", line).group(1)
            # print ("inside name "+l2_vlan_names[L2_VLAN])
        elif re.search("^interface (.*)\s", line):
            VLAN = re.search("^interface (.*)\s", line).group(1)
            #print('Inside interface '+str(VLAN))
        elif re.search("^\s+$", line):
            #print('Ended interface configuration '+str(VLAN)+' '+PORT_MODE+' '+TRUNK)
            # exited from configuration block, VLAN contains the physical interface
            # or the interface vlan.
            if (VLAN):
                if (re.search("Vlan", VLAN)):
                    VLAN_NUM = re.search("Vlan(\d+)", VLAN).group(1)
                if not (re.search("Vlan", VLAN)):
                    VLAN_NUM = 0
                elif ((VLAN_NUM in l2_vlan_names) and re.search("^TR-",l2_vlan_names[VLAN_NUM])):
                    print('Excluded transit '+VLAN+' \"'+l2_vlan_names[VLAN_NUM]+'\"')
                elif (re.search("ansit",description.lower())):
                    print("Excluded transit vlan "+VLAN+" "+description)
                else:
                    if not VRF in if_list[read_file]:
                        if_list[read_file][VRF] = {}
                    if not VLAN in if_list[read_file][VRF]:
                        if_list[read_file][VRF][VLAN] = {}
                    if_list[read_file][VRF][VLAN]['ip address'] = IP_ADDRESS
                # let's save the values on the excel file
                counter += 1
                sheet['A'+str(counter)].value = read_file
                sheet['B'+str(counter)].value = VRF
                # this could be also the physical interface
                sheet['C'+str(counter)].value = VLAN
                if (VLAN_NUM in l2_vlan_names):
                    sheet['D'+str(counter)].value = l2_vlan_names[VLAN_NUM]
                    l2_only_vlans.pop(VLAN_NUM)
                elif (L2_VLAN in l2_vlan_names):
                    sheet['D'+str(counter)].value = l2_vlan_names[L2_VLAN]
                elif (PORT_MODE == 'trunk' and TRUNK!=None):
                    sheet['D'+str(counter)].value = TRUNK
                sheet['E'+str(counter)].value = IP_ADDRESS
                sheet['F'+str(counter)].value = HSRP_GROUP
                sheet['G'+str(counter)].value = HSRP
                sheet['H'+str(counter)].value = description
            # we exited from a configuration 'step', let's erase all the saved parameters
            VLAN = 0
            VRF = ''
            PORT_MODE = ''
            TRUNK = ''
            IP_ADDRESS = ''
            description = ''
            L2_VLAN = 0
            HSRP_GROUP = ''
            HSRP = ''
        # we are inside an interface configuration, being it a physical one or a vlan
        if (VLAN):
            if re.search("^\s+vrf member\s+",line):
                VRF = re.search("vrf member (.*?)\s",line).group(1)
            elif re.search("^\s+description", line):
                description = re.search("^\s+description (.*)\s+",line).group(1)
            elif re.search("ip address", line):
                IP_ADDRESS = re.search("ip address (.*)\s",line).group(1)
            elif re.search("^\s+hsrp\s+\d+", line):
                HSRP_GROUP = re.search("^\s+hsrp\s+(\d+)", line).group(1)
            elif re.search("\s+ip\s+\d+\.\d+\.\d+\.\d+", line):
                mask = re.search("\/(\d+)", IP_ADDRESS).group(1)
                HSRP = re.search("ip\s+(\d+\.\d+\.\d+\.\d+)", line).group(1)+"/"+mask
            elif re.search('switchport access vlan \d+', line):
                L2_VLAN = re.search('switchport access vlan (\d+)', line).group(1)
            elif re.search('switchport access vlan \d+', line):
                L2_VLAN = re.search('switchport access vlan (\d+)', line).group(1)
            elif re.search('switchport mode', line):
                PORT_MODE = re.search('switchport mode (.*?)\s', line).group(1)
            elif re.search('^\s+switchport trunk allowed vlan add \d', line):
                TRUNK += ','+re.search('switchport trunk allowed vlan add (.*?)\s', line).group(1)
            elif re.search('^\s+switchport trunk allowed vlan \d', line):
                TRUNK = re.search('switchport trunk allowed vlan (\d.*?)\s', line).group(1)
    # printing out all L2-only vlans ...
    for vlan in sorted(l2_only_vlans):
        counter += 1
        sheet['A'+str(counter)].value = read_file
        sheet['C'+str(counter)].value = vlan
        sheet['D'+str(counter)].value = l2_vlan_names[vlan]
        sheet['H'+str(counter)].value = 'L2 only vlan'

if 'Vlan_summary' in targ.sheetnames:
    sheet = targ.get_sheet_by_name("Vlan_summary")
else:
    sheet = targ.create_sheet("Vlan_summary")

counter=1
sheet['A'+str(counter)].value = "Apparato"
sheet['B'+str(counter)].value = "vlan"
sheet['C'+str(counter)].value = "vrf"
sheet['D'+str(counter)].value = "#host_ip"
for router in if_list:
    for vrf in if_list[router]:
        counter+=1
        sheet['A'+str(counter)].value = router
        sheet['B'+str(counter)].value = str(len(if_list[router][vrf]))
        sheet['C'+str(counter)].value = vrf
        tot_ip = 0
        for vlan in if_list[router][vrf]:
            if (re.search("\/(\d+)",if_list[router][vrf][vlan]['ip address'])):
                mask = (int)(re.search("\/(\d+)",if_list[router][vrf][vlan]['ip address']).group(1))
                tot_ip += 2**(32 - mask)
        sheet['D'+str(counter)].value = tot_ip

targ.save(excel_dir + file_name)

