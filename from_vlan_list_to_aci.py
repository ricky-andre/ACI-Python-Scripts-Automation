# This python script reads an excel file and generates acitoolkit queries
# to create tenants, bridge-domains, EPG, application profiles and so on.
# The EPG name is strongly related to the BD name (the L2 db vlan name),
# while the application profile where the epg is added could potentially
# be shared, without a 1-to-1 approach ... we do not create by default
# 1 app profile for every epg, the 1-to-1 relationship is only between
# vlans, BD and EPG (even though theoretically a BD could contain MORE vlan).
#
# Given that the bd name is the following one:
# <L2_vlan_name>
# 
# ... in case the app profile name is NOT specified, it will be:
# <vrf_name>_ANP
# 
# ... while the EPG name will be:
# <L2_vlan_name>_EPG

# The excel could be used in more "migration waves", for example we could
# avoid assigning L3 subnets if we just need L2 connectivity, and we could
# add them later on, once all devices have been migrated to the new ACI DC.

# Excel columns can be changed in their appearence order,or new columns can be added
# without the need to modify the script regarding the old columns. The only important
# thing is NOT to change the column's name (if so, perfom a find/replace on this script).
#
# The following checks are performed on the excel data:
#
# 1) check if the tenant already exists
# 2) check if the vrf already exists
# 3) check if the BD already exists
# 4) check if the app profile already exists
# 5) check if the EPG already exists
# 6) print out two tables with the above data
# 7) check if the specified port-channel or vPC interface exist (they should), in the
#     interfaces column (to which EPGs should be added)
# 8) check if the physical interface exists (it should) and is configured, if it is not
#    raise a fatal error.
# 9) check if the tenant, vrf or bd names are empty (they shouldn't)
# 10) if the bd already exists, checks if it has an ip subnet configured, in case
#     in the excel file the subnet is defined, raise a fatal error
# 11) in case the route_type is not specified, it will be "private". In case a value
#     is specified, it must a good one
# 12) in case the EPG name is not specified, the L2 vlan name + "_EPG" will be used
# 13) in case the APP name is not specified, the L2 vlan name + "_ANP" will be used
# 14) succesfully configured objects are green colored, already existing objects are
#     yellow colored, unsuccessful queries are red colored
#
# Beware that ...
# If you use a REST payload with aci_rest using status="created" the action is not idempotent. 
# The first time (when the object does not exist) it works (changed=True), but a subsequent call
# fails as the APIC returns an HTTP failure (i.e. you can't create something that already exists).

# However, if you do the same for removing an object using status="deleted" the action is idempotent.
# So a subsequent call works (and does not report a change, changed=False) as one would expect.
#
# This behaviour, if confirmed on the present APIC's release, is INCONSISTENT ... but for sure
# in my opinion the first is the correct one: if you claim something to be NEW, it should be.
# If you wish to change an object parameters, you can use the "created,modified" status. In this
# case, you will get no errors. 
#
# It should be tested what happens in case of many consecutive queries ... the APIC has a Ddos
# embedded protection mechanisms to limit the number of queries performed by the same user. This
# could slow down the write operations, which are atomically approached for safety reasons:
# a tenant configuration is NOT pushed together with its vrf, bd, epg or whatsoever, operations
# are always performed in a minimal way: a single tenant is created, a single vrf is created and so on.
import openpyxl
from openpyxl.styles import PatternFill
import re
from Aci_Cal_Toolkit import FabLogin, Query, FabTnPol
from credentials import apic_ip, apic_pwd, apic_user
from tabulate import tabulate

# the new changed excel file will be created in the same directory
# 'ACI translate' is the tab name containing all data to be pushed to the APIC
excel_dir = "C:/path_to_excel_data/"
file_name = "ACI_vlan_list.xlsx"
targ = openpyxl.load_workbook(excel_dir + file_name)
sheet = targ.get_sheet_by_name("ACI translate")
xls_cols = {}

# We read the first line of the table, and find the column letter to be assigned to every
# column. In this way, columns can be swapped, new columns can be inserted, and there is
# no need to update the script.
#
# apparato    tenant    vrf    vlan_number    l2_vlan_name    app_profile    ip_addr    description    interfaces
for col in range(0, sheet.max_column):
    xls_cols[sheet[chr(ord('A')+int(col))+'1'].value] = chr(ord('A')+col)
# we will now refer to the columns in the following way:
# xls_cols['variable']
# xls_cols['description']

apic = FabLogin (apic_ip, apic_user, apic_pwd)
cookies = apic.login()
print('Logged into apic ...')

req = Query(apic_ip, cookies)
tnConf = FabTnPol(apic_ip, cookies)

# retrieving all the port-channels, and the full DN path to be added to the EPG
# https://www.cisco.com/c/en/us/td/docs/switches/datacenter/aci/apic/sw/2-x/rest_cfg/2_1_x/b_Cisco_APIC_REST_API_Configuration_Guide/b_Cisco_APIC_REST_API_Configuration_Guide_chapter_01.html
#
# check out here the possibilities to perform queries, available filters and so on ...
# query_url = ('/api/mo/topology/.json?query-target=subtree&target-subtree-class=fabricProtPathEpCont')
#
# in this case quering for fabricPathEp return 1900 objects ... too many
# so we need to keep this way of doing it, even though it requires a few queries
# and it is time expensive.
print('Retrieving all the port-channels and vPC ...')
vpc_dn = req.query_vpc()

print('Retrieving all fabric ports information ...')
ports_data = req.query_ports()

# let's read all the present tenants, vrf, BDs to check for double information
# and already existent policies.
print("Retrieving all tenants information (vrf, bd, subnets, anp, epg) could take a while ...\n\n")
apic_data = req.query_all_tenants()


fatal_error = 0
# we now check that if BD are NOT new, we are not over-writing something. In case
# the ip address is defined in the excel, the BD already exists but it doesn't
# have an ip defined, we just print a warning. If an ip address already exists,
# we raise a fatal error: we suppose here that every BD can be associated to just
# one subnet. Regadring application profile, we do not think there are problems
# in case the application profile already exists or not. A warning could be printed
# anyway in case the application profile already exists (????).
# We also check that in every row a tenant and a vrf is defined.
print('Checking for errors inside the excel file on tenant, vrf, BD columns ...')
for row in range(2, sheet.max_row+1):
    bd_name = str(sheet[xls_cols['l2_vlan_name']+str(row)].value).strip()
    if bd_name == 'None':
        print ('ERROR in row '+str(row)+' empty BD name')
        fatal_error = 1
        continue
    tenant = str(sheet[xls_cols['tenant']+str(row)].value).strip()
    if tenant == 'None':
        print ('ERROR in row '+str(row)+' empty TENANT name')
        fatal_error = 1
        continue
    vrf = str(sheet[xls_cols['vrf']+str(row)].value).strip()
    if vrf == 'None':
        print ('ERROR in row '+str(row)+' empty VRF name')
        fatal_error = 1
        continue
    sub_scope = str(sheet[xls_cols['route_type']+str(row)].value).strip()
    if (sheet[xls_cols['route_type']+str(row)].value != None):
        sub_scope = str(sheet[xls_cols['route_type']+str(row)].value).strip()
        if (sub_scope!='public' and  sub_scope!='private' and sub_scope!='shared'):
            print ('ERROR in row '+str(row)+' route-type must be private|public|shared')
            fatal_error = 1
            continue
    ip_addr = str(sheet[xls_cols['ip_addr']+str(row)].value).strip()
    if tenant in apic_data:
        if bd_name in apic_data[tenant]['bd_list']:
            if len(apic_data[tenant]['bd_list'][bd_name]['ip']):
                print ('ERROR in row '+str(row)+' BD and subnet(s) already exists !')
                fatal_error = 1
                continue
            else:
                print ('WARNING in row '+str(row)+' BD already exists without a configured subnet')
            if not 'vrf' in apic_data[tenant]['bd_list'][bd_name]:
                print ('WARNING in row '+str(row)+' BD already exists without a configured vrf')

if (fatal_error):
    print ("Exiting, fatal errors occurred on excel data")
    exit(0)

# we now check the vlan columns and the consistency of the interfaces columns
fatal_error = 0
print('Checking interfaces column, this could take a while ...')
for row in range(2, sheet.max_row+1):
    # checking for Vlan number consistency
    if not re.search("Vlan\d+",sheet[xls_cols['vlan_number']+str(row)].value):
        print ('ERROR data on row '+str(row)+' wrong vlan number value, must be Vlan<id>')
        fatal_error = False
    vlan_id = re.search("Vlan(\d+)",sheet[xls_cols['vlan_number']+str(row)].value).group(1)
    if (str(sheet[xls_cols['interfaces']+str(row)].value).strip() != 'None'):
        # let's retrieve all the lines of the interfaces
        lines = str(sheet[xls_cols['interfaces']+str(row)].value).strip().splitlines()
        xls_port = None
        for intf in lines:
            if re.search(",", intf):
                if_error = 0
                data = intf.split(",")
                if len(data) != 5:
                    fatal_error = True
                    if_error = 1
                    print ('ERROR on row '+str(row)+', must be vPC or format \'eth,pod,node_id,module,port_id\'')
                for i in range(1,5):
                    if not re.search("\d+", data[i]):
                        fatal_error = True
                        if_error = 1
                        print ('ERROR on row '+str(row)+', must be vPC or format \'eth,pod,node_id,module,port_id\'')
                node_id = (int)(data[2])
                port = data[3]+"/"+data[4]
                # check that the interface is NOT part of a bundle. In this case print an error and
                # the "policy group", raise a fatal error and stop. This must be manually solved,
                # the interface could be wrong.
                if node_id in ports_data:
                    if port in ports_data[node_id]['ports']:
                        # this means that the interface is used, there is a policy applied
                        # over it, otherwise it would not be here.
                        if not ports_data[node_id]['ports'][port]['type']=='access':
                            fatal_error = True
                            print('ERROR on row '+str(row)+' port is in a bundle, use polGrp name '+
                                  ports_data[node_id]['ports'][port]['polGrp']+' or change port')
                    else:
                        fatal_error = True
                        print('ERROR on row '+str(row)+' port not configured '+intf)
                else:
                    fatal_error = True
                    print('ERROR on row '+str(row)+' selected node does not exist')
                
                if not if_error:
                    query_url = ('/api/mo/topology/pod-%s/node-%s.json?query-target=subtree'
                     '&target-subtree-class=l1PhysIf' % (data[1], data[2]))
                    query_url += '&query-target-filter=eq(l1PhysIf.id,"%s")' % (data[0]+data[3]+"/"+data[4])
                    # print (query_url)
                    [status, payload] = req.query_url(query_url)                    
                    json_data = payload['imdata']
                    if not len(json_data)>0:
                        fatal_error = True
                        print ('ERROR on row '+str(row)+', wrong interface data value '+intf)
                xls_port ='eth'+data[3]+"/"+data[4]
            else:
                # in this case it MUST be a port channel name
                if not intf.strip() in vpc_dn:
                    fatal_error = True
                    print ('ERROR on row '+str(row)+' for vpc name, should be one of the following:')
                    for key in vpc_dn:
                        print (key)
                xls_port = intf.strip()
        # Here it should also be checked if on the port/vPC, the vlan assigned to the EPG
        # is free or is used. In case it is used, print the EPG that uses it. This check
        # could give some false positive, in case the same policy group is used on
        # many different bundles. I believe this is possible in ACI, in case there are different
        # interface selectors that match different ranges using the same policy (?????).
        if not fatal_error:
            [status, payload] = req.query_url('/api/node/class/fvRsPathAtt.json?query-target-filter=eq(fvRsPathAtt.encap,"vlan-'+vlan_id+'")')
            json_data = payload['imdata']
            for obj in json_data:
                dn = obj['fvRsPathAtt']['attributes']['dn']
                reg = re.search('\/epg-(.*?)\/.*\/pathep-\[(.*?)\]', dn)
                epg = reg.group(1).strip()
                port = reg.group(2).strip()
                if port == xls_port:
                    fatal_error = True
                    print ('ERROR on row '+str(row)+' encap vlan already used by '+epg)

if (fatal_error):
    print ("Exiting, fatal errors occurred on excel data")
    exit(0)

# Reading all the tenant, vrf, app profiles used in the excel file, checking if they 
# are new or not and printing it out. For already existing tenants,
# we also print out if vrf already exist. It is important that we don't change
# anything important !! For example the 'enforced/unenforced' attribute, which could
# break the whole vrf connectivity ... with AciToolKit, by default a vrf json data
# was created with the 'enforced' attribute set.
xls_data = {}
xls_app = {}
for row in range(2, sheet.max_row+1):
    ten_name = str(sheet[xls_cols['tenant']+str(row)].value).strip()
    vrf_name = str(sheet[xls_cols['vrf']+str(row)].value).strip()
    app_name = str(sheet[xls_cols['app_profile']+str(row)].value).strip()
    bd_name = str(sheet[xls_cols['l2_vlan_name']+str(row)].value).strip()+"_BD"
    if (app_name == 'None'):
        app_name = vrf_name + "_ANP"
    epg_name = str(sheet[xls_cols['l2_vlan_name']+str(row)].value).strip()+"_EPG"
    
    if not ten_name in xls_data:
        xls_data[ten_name] = {}
        xls_app[ten_name] = {}
    if not vrf_name in xls_data[ten_name]:
        xls_data[ten_name][vrf_name] = {}
    if not bd_name in xls_data[ten_name][vrf_name]:
        xls_data[ten_name][vrf_name][bd_name]=1
    if not app_name in xls_app[ten_name]:
        xls_app[ten_name][app_name] = {}
    if not epg_name in xls_app[ten_name][app_name]:
        xls_app[ten_name][app_name][epg_name] = 1

output_data = []
for ten_name in xls_data:
    for vrf_name in xls_data[ten_name]:
        for bd in xls_data[ten_name][vrf_name]:
            if not ten_name in apic_data:
                output_data.append((ten_name, 'new', vrf_name, 'new', bd, 'new'))
            elif not vrf_name in apic_data[ten_name]['vrf_list']:
                output_data.append((ten_name, 'exists', vrf_name, 'new', bd, 'new'))
            elif not bd in apic_data[ten_name]['vrf_list'][vrf_name]:
                output_data.append((ten_name, 'exists', vrf_name, 'exists', bd, 'new'))
            else:
                output_data.append((ten_name, 'exists', vrf_name, 'exists', bd, 'exists'))
print (tabulate(output_data, headers=["tenant", "new", "vrf", "new", "BD", "new"]))
print ('\n\n')

output_data = []
for ten_name in xls_app:
    for app in xls_app[ten_name]:
        for epg in xls_app[ten_name][app]:
            if not ten_name in apic_data:
                output_data.append((ten_name, 'new', app, 'new', epg, 'new'))
            elif not app in apic_data[ten_name]['anp_list']:
                output_data.append((ten_name, 'exists', app, 'new', epg, 'new'))
            elif not epg in apic_data[ten_name]['anp_list'][app]:
                output_data.append((ten_name, 'exists', app, 'exists', epg, 'new'))
            else:
                output_data.append((ten_name, 'exists', app, 'exists', epg, 'exists'))
print (tabulate(output_data, headers=["tenant", "new", "app profile", "new", "epg", "new"]))


def color (status, cell):
    if status == None:
        sheet[cell].fill = PatternFill("solid", fgColor="FFFF00")
    elif status == 200:
        sheet[cell].fill = PatternFill("solid", fgColor="00FF00")
    else:
        sheet[cell].fill = PatternFill("solid", fgColor="FF1100")        

print('\n\nCreating queries, row by row ...')
# To keep everything more simple and clean, it is better to perform a query
# row by row. Objects are created in the logical order:
# - tenant
# - vrf
# - bridge domains (with subnets)
# - application profiles
# - end point groups
# 
# In case the object creation is successful, the cell is green colored, and the
# object is added to the structures. In this way, the second time the same object
# should be created, the operation is simply SKIPPED and the cell is yellow colored.
# In case of errors, the cell is red colored and some other object creation could
# be skipped. For example in case a vrf can't be created, it doesn't make any sense 
# to create a bridge domain that points to that vrf.
# yellow FFFF00
# red FF1100
# green 00FF00
for row in range(2, len(sheet['A'])+1):
    # TENANT
    ten_name = str(sheet[xls_cols['tenant']+str(row)].value).strip()
    status = None
    if not ten_name in apic_data:
        status = tnConf.tenant(name = ten_name, 
                               status = 'created')
        if status == 200:
            apic_data[ten_name]={}
            apic_data[ten_name]['vrf_list'] = {}
            apic_data[ten_name]['anp_list'] = {}
            apic_data[ten_name]['bd_list'] = {}
            apic_data[ten_name]['anp_list'] = {}
    color(status, xls_cols['tenant']+str(row))
    
    # VRF
    vrf_name = str(sheet[xls_cols['vrf']+str(row)].value).strip()
    status = None
    if not vrf_name in apic_data[ten_name]['vrf_list']:
        status = tnConf.vrf(tn_name = ten_name,
                            name = vrf_name,
                            enforce = 'enforced',
                            status = 'created')
        if status == 200:
            apic_data[ten_name]['vrf_list'][vrf_name]={}
    color(status, xls_cols['vrf']+str(row))
    
    l2_vname = str(sheet[xls_cols['l2_vlan_name']+str(row)].value).strip()
    vlan_num = re.search("Vlan(\d+)",sheet[xls_cols['vlan_number']+str(row)].value).group(1)
    description = sheet[xls_cols['descr']+str(row)].value.strip()
    
    sub_name = None
    sub_addr = None
    sub_scope = 'private'
    if (sheet[xls_cols['ip_addr']+str(row)].value != None):
        sub_name = l2_vname + '_subnet'
        sub_addr = sheet[xls_cols['ip_addr']+str(row)].value.strip()
    # we have already checked that there are only right values
    if (sheet[xls_cols['route_type']+str(row)].value != None):
        sub_scope = str(sheet[xls_cols['route_type']+str(row)].value).strip()
    else:
        sheet[xls_cols['route_type']+str(row)].value = 'private'
    
    # BRIDGE DOMAIN
    # bd parameters should be optimized depending on other data.
    bd_name = l2_vname + "_BD"
    status = None
    if not bd_name in apic_data[ten_name]['vrf_list'][vrf_name]:
        if (sub_addr == None):
            # configurations optimized for a L2-only bridge domain, changing the 
            # parameters later when a BD is added could be intrusive, for example
            # changing the approach of unk_unicast to spine-proxy would lead to
            # some packet loss. There is no subnet configured, thus unicast routing
            # MUST be disabled, and arp flooding must be enabled.
            status = tnConf.bd(tn_name = ten_name,
                               name = bd_name,
                               arp = 'yes',
                               mdest = 'bd-flood',
                               mcast = 'flood',
                               unicast = 'no',
                               unk_unicast = 'flood',
                               status = 'created',
                               vrf = vrf_name,
                               descr = description)
        else:
            status = tnConf.bd(tn_name = ten_name,
                               name = bd_name,
                               arp = 'no',
                               mdest = 'bd-flood',      # ???
                               mcast = 'opt-flood',     # ???
                               unicast = 'yes',
                               unk_unicast = 'proxy',
                               status = 'created',
                               limitlearn = 'yes',
                               vrf = vrf_name,
                               descr = description)
        
        if (status == 200):
            apic_data[ten_name]['bd_list'][bd_name]={}
        
        if (status == 200):
            # add the vrf to the BD
            status_vrf = tnConf.bd_vrf(tn_name = ten_name,
                                       name = bd_name,
                                       status = 'created',
                                       vrf = vrf_name)
            if (status_vrf == 200):
                apic_data[ten_name]['vrf_list'][vrf_name][bd_name]={}
        
        # in case of errors creating the new BD or the vrf associated to it,
        # the status value is signed as wrong. The post getting the error will
        # be printed out on screen.        
        if (status == 200 and status_vrf==200):
            status=200
        else:
            status = 0
    color(status, xls_cols['l2_vlan_name']+str(row))
    
    # in case the BD already exists, we suppose it has already been associated to a vrf,
    # since EVERY bd must belong to a vrf in ACI. In case it's not
    if not bd_name in apic_data[ten_name]['vrf_list'][vrf_name]:
        status_vrf = tnConf.bd_vrf(tn_name = ten_name,
                                       name = bd_name,
                                       status = 'created',
                                       vrf = vrf_name)
        if (status_vrf == 200):
            apic_data[ten_name]['vrf_list'][vrf_name][bd_name]={}
        print('Warning, added '+vrf_name+' to BD '+bd_name)
    
    # BD SUBNET
    # if the BD already exists, the L3 subnet should be added, it was already
    # checked if another subnet was already there ... in this case a fatal error
    # would have been printed out, and the script would have blocked itself.
    status = tnConf.bd_subnet(tn_name = ten_name,
                              name = bd_name,
                              subnet = sub_addr,
                              scope = sub_scope,
                              preferred = 'yes',
                              status = 'created',
                              descr = description)
    if status == 200:
        # add the NEW subnet to the BD
        apic_data[ten_name]['vrf_list'][vrf_name][bd_name]['ip'] = []
        apic_data[ten_name]['bd_list'][bd_name]['ip'] = []
        
        apic_data[ten_name]['vrf_list'][vrf_name][bd_name]['ip'].append(sub_addr)
        apic_data[ten_name]['bd_list'][bd_name]['ip'].append(sub_addr)
    color(status, xls_cols['ip_addr']+str(row))
    
    
    if (sheet[xls_cols['app_profile']+str(row)].value != None):
        app_prof = str(sheet[xls_cols['app_profile']+str(row)].value).strip()
    else:
        app_prof = vrf_name + "_ANP"
        sheet[xls_cols['app_profile']+str(row)].value = app_prof
    
    # APP PROFILE
    status = None
    if not app_prof in apic_data[ten_name]['anp_list']:
        status = tnConf.app_profile(tn_name = ten_name,
                                    name = app_prof,
                                    status = 'created')
        if status == 200:
            apic_data[ten_name]['anp_list'][app_prof]={}
    color(status, xls_cols['app_profile']+str(row))
    
    # EPG
    if (sheet[xls_cols['epg']+str(row)].value != None):
        epg = str(sheet[xls_cols['epg']+str(row)].value).strip()
    else:
        epg = l2_vname+"_EPG"
        sheet[xls_cols['epg']+str(row)].value = epg
    status = None
    if not epg in apic_data[ten_name]['anp_list'][app_prof]:
        status = tnConf.epg(tn_name = ten_name,
                            ap_name = app_prof,
                            name = epg,
                            bd = bd_name,
                            status = 'created')
        if status == 200:
            apic_data[ten_name]['anp_list'][app_prof][epg] = {}
    color(status, xls_cols['epg']+str(row))
    
    # the domain is always the same, we define it here anyway. In general,
    # it could be different for every port but usually this is what is done:
    # it is created a physical domain which contains all the vlan except the intra-vlan,
    # for SNAM it is this one: "Physical_PhysDom"
    #
    # Remember that we have already checked that the port is configured, the vPC exists,
    # and that encap vlan on the physical interface is NOT already used.
    multiple_checks = True
    if (str(sheet[xls_cols['interfaces']+str(row)].value).strip() != 'None'):
        lines = str(sheet[xls_cols['interfaces']+str(row)].value).strip().splitlines()
        for intf in lines:
            if re.search(",",intf):
                # 'eth', pod, leaf-id, module, port
                if_data = intf.split(",")
                status = tnConf.static_path_access(tn_name = ten_name,
                                                ap_name = app_prof,
                                                epg_name = epg,
                                                sw1 = if_data[2],
                                                port = if_data[4],
                                                encap = vlan_num,
                                                deploy = 'immediate',
                                                status = 'created',
                                                pod = if_data[1])
                if status != 200:
                    multiple_checks = False
            else:
                # it must be a vPC or a port-channel
                # topology/pod-1/protpaths-113-114/pathep-[vPC_CRE-SNCLB00171_PolGrp]
                res = re.search('\/pod-(\d+)\/protpaths-(\d+)-(\d+)\/pathep-\[(.*)\]', vpc_dn[intf.strip()])
                status = tnConf.static_path_vpc(tn_name = ten_name,
                                                ap_name = app_prof,
                                                epg_name = epg,
                                                sw1 = res.group(2),
                                                sw2 = res.group(3),
                                                vpc = res.group(4),
                                                encap = vlan_num,
                                                deploy = 'immediate',
                                                status = 'created',
                                                pod = res.group(1))
                if status != 200:
                    multiple_checks = False
    if multiple_checks:
        status = 200
    if (sheet[xls_cols['interfaces']+str(row)].value != None):
        color(status, xls_cols['interfaces']+str(row))

targ.save('C:/Users/601787621/Documenti/Snam/ACI e VMWare/ACI_vlan_list_out.xlsx')
