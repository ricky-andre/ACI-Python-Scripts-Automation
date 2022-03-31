# This python script retrives data from ACI and saves them into an excel
# for an easier processing and filtering capabilities for everyone.
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
import re
import time
import os.path
from Aci_Cal_Toolkit import FabLogin, Query
from credentials import apic_ip, apic_pwd, apic_user

# clock_time = time.strftime("%H %M %S")
day_time = time.strftime("a%Y m%m g%d")
if os.path.exists('C:/<path_to_excel_output_dir>/ACI_class_' + day_time + '.xlsx'):
    target = load_workbook('C:/<path_to_excel_output_dir>/ACI_class_' + day_time + ".xlsx")
else:
    target = Workbook()

apic = FabLogin (apic_ip, apic_user, apic_pwd)
cookies = apic.login()
print('Logged into apic ...')
req = Query(apic_ip, cookies)

fvClass = ['mgmtRsInBStNode', 'fvTenant']

# for bridge domains, it would be interesting to know which have an ip subnet associated
# to them, and which of them do not. This is helpful to make comparisons between all the
# flooding configurations (but not limited to them).
if 'fvBD' in fvClass:
    bd_nets = {}
    print('Retrieving BD subnets ...')
    [status, payload] = req.query_class('fvSubnet')
    if (status != 200):
        print('Error retrieving fvSubnet')
        exit(0)
    json_data = payload['imdata']
    for obj in json_data:
        dn = obj['fvSubnet']['attributes']['dn']
        # uni/tn-<tn_name>/BD-<bd_name>/subnet-[<subnet>]
        # There are also the following objects in case subnets are created
        # as 'sons' of epg, this MUST be done in case contracts are created
        # between different vrf through route-leaking:
        #  uni/tn-<tn_name>/ap-<anp_name>/epg-<epg_name>/subnet-[<subnet>]
        reg = re.search('\/tn-(.*?)\/BD-(.*?)\/subnet-\[(.*)\]', dn)
        if reg != None:
            ten_name = reg.group(1)
            bd_name = reg.group(2)
            ip = reg.group(3)
            if not bd_name in bd_nets:
                bd_nets[bd_name] = []
            bd_nets[bd_name].append(ip)


for aci_class in fvClass:
    sheet = target.create_sheet(aci_class, 0)
    print('Retrieving class '+aci_class+' ...')
    [status, payload] = req.query_class(aci_class)
    if status != 200 or not len(payload['imdata']):
        continue
    json_data = payload['imdata']
    cols = []
    for attr in sorted(json_data[0][aci_class]['attributes']):
        cols.append(attr)
        sheet[get_column_letter(len(cols))+'1'] = attr
    row_counter = 2
    for obj in json_data:
        for i in range(len(cols)):
            sheet[get_column_letter(i+1)+str(row_counter)] = obj[aci_class]['attributes'][cols[i]]
        if aci_class == 'fvBD' and (obj[aci_class]['attributes']['name'] in bd_nets):
            nets = '\n'.join(bd_nets[obj[aci_class]['attributes']['name']])
            sheet[get_column_letter(len(cols)+1)+str(row_counter)] = nets
        row_counter += 1

target.save('C:/<path_to_excel_output_dir>/ACI_class_'+day_time+".xlsx")
